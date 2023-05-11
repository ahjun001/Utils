#!/usr/bin/env python3
"""
tranlate excel column in English to French
"""
from time import sleep, time
from googletrans import Translator   # googletrans==3.1.0a0
import openpyxl    # openpyxl==3.1.2

# timing overall program, pauses aside: 1.5 sec / translation
start_time = time()


def translate(anglais):
    """
    # Function to translate an English word or sentence to French
    """

    # Get the Google Translate API client
    client = Translator()

    # Translate the English word or sentence to French & return
    return client.translate(anglais, dest="fr", src="en").text


# Get the Excel spreadsheet
MY_EXCEL_FILE = "Translated pairs 1 sheet.xlsx"
wb = openpyxl.load_workbook(MY_EXCEL_FILE)

# Get the sheet with the English words and sentences
sheet = wb["For proofreading"]

# Loop through the English words and sentences
for cell in sheet['C']:

    if cell.row % 3 == 0:
        wb.save(MY_EXCEL_FILE)

    if cell.row % 10 == 0:
        sleep(10)

    if cell.row % 105 == 0:
        sleep(10)

    if not cell.value:
        continue

    if cell.row > sheet.max_row:
        break

    # Get the English word or sentence
    english = cell.value
    print(english, end=' -- ')

    # Translate the English word or sentence to French
    french = translate(english)
    french = french or 'n/a'
    print(french)

    # Write the French translation to the new column
    sheet.cell(row=cell.row, column=9).value = french

# Save the Excel spreadsheet
wb.save(MY_EXCEL_FILE)

print(f"\n--- {time() - start_time:.2f} seconds ---")
